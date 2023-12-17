import os
import shutil
from openpyxl import load_workbook

parent_directory = "/Users/josetiznado/Documents"
base_directory = "/Users/josetiznado/Desktop/Job_Projects"
target_directory = ""

def get_target_directory(file):
    file_prefix = str(file)[:4]
    return f'{base_directory}/{file_prefix}/Purchasing/MTRs'

def move_file(file, file_path):
    global target_directory
    target_directory = get_target_directory(file)
    
    if os.path.isdir(target_directory):
        os.chdir(target_directory)

    shutil.move(file_path, target_directory)

def create_hyperlink(file_path, file_name):
    spreadsheet = f'{os.path.dirname(target_directory)}/Hyperlinks.xlsx'
    wb = load_workbook(spreadsheet)
    sheet = wb.active

    stem = os.path.splitext(file_name)[0]
    for cell in sheet['A']:
        if any([stem.endswith("-1"),
                stem.endswith("-2"),
                stem.endswith("-3"),
                stem.endswith("-4")]):
            cell_value = sheet.cell(row = cell.row, column = cell.column + int(stem[-1]) - 1).value
            if os.path.splitext(str(cell.value))[0] == stem[:-2]:
                sheet.cell(row = cell.row, column = cell.column + int(stem[-1]) - 1).value = '=HYPERLINK("{}", "{}")'.format(file_path, cell_value)

        elif cell.value == stem:
            sheet.cell(row = cell.row, column = cell.column).value = '=HYPERLINK("{}", "{}")'.format(file_path, file_name)

    wb.save(spreadsheet)

if __name__ == '__main__':
    files = os.listdir(parent_directory)

    for file in files:
        if file == ".DS_Store":
            continue

        file_path = f'{parent_directory}/{file}'

        if os.path.isdir(file_path):
            continue

        move_file(file, file_path)

        new_file_path = f'{target_directory}/{file}'

        create_hyperlink(new_file_path, file)